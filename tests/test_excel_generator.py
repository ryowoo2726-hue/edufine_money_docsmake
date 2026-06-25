from decimal import Decimal

from openpyxl import load_workbook
from openpyxl import Workbook

from auto_money_doc.excel.generator import generate_excel
from auto_money_doc.excel.template_mapper import infer_template_mapping
from auto_money_doc.extraction.schema import ApprovalMetadata, QuotationData, QuotationItem


def test_generate_excel_uses_source_file_name(tmp_path):
    quotation = QuotationData(
        source_file_names=["sample_quote.pdf"],
        vendor_name="테스트상사",
        total_amount=Decimal("12000"),
        items=[
            QuotationItem(
                item_name="복사용지",
                quantity=Decimal("2"),
                unit="박스",
                unit_price=Decimal("6000"),
                total_amount=Decimal("12000"),
            )
        ],
    )

    output = generate_excel(
        quotation=quotation,
        approval_metadata=ApprovalMetadata(school_name="테스트중학교"),
        output_path=str(tmp_path),
    )

    assert output.name == "sample_quote_품의서.xlsx"
    assert output.exists()

    workbook = load_workbook(output)
    sheet = workbook["품의서"]
    assert [sheet.cell(row=1, column=col).value for col in range(1, 7)] == [
        "내용",
        "규격",
        "수량",
        "단위",
        "예상단가",
        "예상금액",
    ]
    assert sheet["A2"].value == "복사용지"
    assert sheet["C2"].value == 2
    assert sheet["D2"].value == "박스"
    assert sheet["E2"].value == 6000
    assert sheet["F2"].value == 12000


def test_infer_mapping_and_generate_with_template(tmp_path):
    template_path = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "품의서"
    sheet["A1"] = "학교명"
    sheet["A2"] = "업체명"
    sheet["A3"] = "합계"
    sheet["A5"] = "품명"
    sheet["B5"] = "규격"
    sheet["C5"] = "수량"
    sheet["D5"] = "단가"
    sheet["E5"] = "합계금액"
    workbook.save(template_path)

    mapping = infer_template_mapping(str(template_path)).template_mapping
    quotation = QuotationData(
        source_file_names=["quote_a.pdf", "quote_b.pdf"],
        vendor_name="미래문구",
        total_amount=Decimal("25000"),
        items=[
            QuotationItem(
                item_name="보드마카",
                specification="검정",
                quantity=Decimal("10"),
                unit_price=Decimal("2500"),
                total_amount=Decimal("25000"),
            )
        ],
    )

    output = generate_excel(
        quotation=quotation,
        approval_metadata=ApprovalMetadata(school_name="테스트중학교"),
        template_path=str(template_path),
        output_path=str(tmp_path / "result.xlsx"),
        template_mapping=mapping,
    )

    result = load_workbook(output)
    sheet = result["품의서"]
    assert sheet["B1"].value == "테스트중학교"
    assert sheet["B2"].value == "미래문구"
    assert sheet["B3"].value == 25000
    assert sheet["A6"].value == "보드마카"
    assert sheet["E6"].value == 25000

    summary = result["자동추출데이터"]
    assert summary["B9"].value == "미래문구"
