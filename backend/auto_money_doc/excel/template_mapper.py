from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from auto_money_doc.extraction.schema import (
    CellMapping,
    InferTemplateMappingResponse,
    ItemTableMapping,
    TemplateMapping,
)


SCALAR_FIELD_LABELS = {
    "school_name": ["학교명", "학교"],
    "department": ["부서", "담당부서", "소속"],
    "requester": ["작성자", "기안자", "담당자"],
    "budget_category": ["예산과목", "예산", "세부사업"],
    "project_name": ["사업명", "건명", "제목"],
    "purchase_purpose": ["구입목적", "구매목적", "목적", "용도"],
    "request_date": ["요청일", "작성일", "기안일", "구입일"],
    "vendor_name": ["업체명", "상호", "거래처", "공급자"],
    "vendor_business_number": ["사업자번호", "등록번호"],
    "vendor_phone_number": ["전화번호", "연락처", "전화"],
    "quotation_date": ["견적일", "견적일자"],
    "validity_period": ["유효기간"],
    "contact_person": ["업체담당자", "담당자"],
    "supply_amount": ["공급가액", "공급액"],
    "tax_amount": ["부가세", "세액", "VAT"],
    "total_amount": ["합계금액", "총액", "합계", "총금액"],
    "notes": ["비고", "특이사항"],
}

ITEM_FIELD_LABELS = {
    "item_name": ["품명", "물품명", "내역", "품목"],
    "specification": ["규격", "모델", "사양"],
    "quantity": ["수량"],
    "unit": ["단위"],
    "unit_price": ["단가"],
    "supply_amount": ["공급가액", "공급액", "금액"],
    "tax_amount": ["세액", "부가세", "VAT"],
    "total_amount": ["합계", "합계금액", "금액"],
    "notes": ["비고"],
}


def _normalize(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _label_score(cell_value: object, labels: list[str]) -> tuple[float, str]:
    normalized = _normalize(cell_value)
    if not normalized:
        return 0.0, ""

    for label in labels:
        normalized_label = _normalize(label)
        if normalized == normalized_label:
            return 1.0, label
        if normalized_label in normalized:
            return 0.82, label
    return 0.0, ""


def _target_cell_for_label(cell: Cell) -> str:
    ws = cell.parent
    right = ws.cell(row=cell.row, column=cell.column + 1)
    if right.value in (None, ""):
        return right.coordinate

    below = ws.cell(row=cell.row + 1, column=cell.column)
    if below.value in (None, ""):
        return below.coordinate

    return right.coordinate


def _infer_scalar_fields(workbook) -> list[CellMapping]:
    found: dict[str, CellMapping] = {}
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                for field_key, labels in SCALAR_FIELD_LABELS.items():
                    score, label = _label_score(cell.value, labels)
                    previous = found.get(field_key)
                    if score and (previous is None or score > previous.confidence):
                        found[field_key] = CellMapping(
                            field_key=field_key,
                            sheet_name=ws.title,
                            cell=_target_cell_for_label(cell),
                            confidence=score,
                            source_label=label,
                        )
    return list(found.values())


def _infer_item_table(workbook) -> ItemTableMapping | None:
    best_mapping: ItemTableMapping | None = None
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            columns: dict[str, str] = {}
            score_sum = 0.0
            for cell in row:
                for field_key, labels in ITEM_FIELD_LABELS.items():
                    score, _ = _label_score(cell.value, labels)
                    if score and field_key not in columns:
                        columns[field_key] = get_column_letter(cell.column)
                        score_sum += score

            if len(columns) < 3 or "item_name" not in columns:
                continue

            confidence = min(score_sum / max(len(ITEM_FIELD_LABELS), 1), 1.0)
            candidate = ItemTableMapping(
                sheet_name=ws.title,
                start_row=row[0].row + 1,
                columns=columns,
                confidence=confidence,
            )
            if best_mapping is None or candidate.confidence > best_mapping.confidence:
                best_mapping = candidate

    return best_mapping


def infer_template_mapping(template_path: str) -> InferTemplateMappingResponse:
    template = Path(template_path).expanduser()
    if not template.exists():
        raise FileNotFoundError(f"Excel template does not exist: {template}")

    workbook = load_workbook(template)
    scalar_fields = _infer_scalar_fields(workbook)
    item_table = _infer_item_table(workbook)
    warnings: list[str] = []

    if not scalar_fields:
        warnings.append("No scalar fields were inferred from the template.")
    if not item_table:
        warnings.append("No item table was inferred from the template.")

    return InferTemplateMappingResponse(
        template_mapping=TemplateMapping(
            scalar_fields=scalar_fields,
            item_table=item_table,
            warnings=warnings,
        )
    )
