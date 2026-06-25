from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from auto_money_doc.extraction.schema import (
    ApprovalMetadata,
    QuotationData,
    TemplateMapping,
)


INVALID_FILENAME_CHARS = r'[<>:"/\\|?*]'


def safe_filename(value: str, fallback: str = "quotation") -> str:
    name = re.sub(INVALID_FILENAME_CHARS, "_", value).strip().strip(".")
    return name or fallback


def build_default_output_path(quotation: QuotationData, output_path: str) -> Path:
    if output_path:
        path = Path(output_path).expanduser()
        if path.suffix.lower() == ".xlsx":
            return path
        source_name = quotation.source_file_names[0] if quotation.source_file_names else ""
        stem = safe_filename(Path(source_name).stem, "quotation")
        return path / f"{stem}_품의서.xlsx"

    source_name = quotation.source_file_names[0] if quotation.source_file_names else ""
    stem = safe_filename(Path(source_name).stem, "quotation")
    return Path.cwd() / "outputs" / f"{stem}_품의서.xlsx"


def _append_summary_sheet(
    workbook: Workbook,
    quotation: QuotationData,
    approval_metadata: ApprovalMetadata,
) -> None:
    sheet_name = "자동추출데이터"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name, 0)

    rows = [
        ("생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("학교명", approval_metadata.school_name),
        ("부서", approval_metadata.department),
        ("작성자", approval_metadata.requester),
        ("예산과목", approval_metadata.budget_category),
        ("사업명", approval_metadata.project_name),
        ("구입목적", approval_metadata.purchase_purpose),
        ("요청일", approval_metadata.request_date),
        ("업체명", quotation.vendor_name),
        ("사업자번호", quotation.vendor_business_number),
        ("업체전화", quotation.vendor_phone_number),
        ("견적일", quotation.quotation_date),
        ("유효기간", quotation.validity_period),
        ("담당자", quotation.contact_person),
        ("공급가액", quotation.supply_amount),
        ("세액", quotation.tax_amount),
        ("합계", quotation.total_amount),
        ("비고", quotation.notes),
    ]

    for index, (label, value) in enumerate(rows, start=1):
        ws.cell(row=index, column=1, value=label)
        ws.cell(row=index, column=2, value=value)

    start_row = len(rows) + 3
    headers = ["품명", "규격", "수량", "단위", "단가", "공급가액", "세액", "합계", "비고"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=header)

    for row_offset, item in enumerate(quotation.items, start=1):
        row = start_row + row_offset
        values = [
            item.item_name,
            item.specification,
            item.quantity,
            item.unit,
            item.unit_price,
            item.supply_amount,
            item.tax_amount,
            item.total_amount,
            item.notes,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18


def _populate_default_approval_sheet(workbook: Workbook, quotation: QuotationData) -> None:
    ws = workbook.active
    ws.title = "품의서"

    headers = ["내용", "규격", "수량", "단위", "예상단가", "예상금액"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_offset, item in enumerate(quotation.items, start=2):
        ws.cell(row=row_offset, column=1, value=item.item_name)
        ws.cell(row=row_offset, column=2, value=item.specification)
        ws.cell(row=row_offset, column=3, value=item.quantity)
        ws.cell(row=row_offset, column=4, value=item.unit)
        ws.cell(row=row_offset, column=5, value=item.unit_price)
        ws.cell(row=row_offset, column=6, value=item.total_amount)

    widths = {
        "A": 34,
        "B": 24,
        "C": 12,
        "D": 12,
        "E": 16,
        "F": 16,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _field_values(
    quotation: QuotationData,
    approval_metadata: ApprovalMetadata,
) -> dict[str, object]:
    return {
        "school_name": approval_metadata.school_name,
        "department": approval_metadata.department,
        "requester": approval_metadata.requester,
        "budget_category": approval_metadata.budget_category,
        "project_name": approval_metadata.project_name,
        "purchase_purpose": approval_metadata.purchase_purpose,
        "request_date": approval_metadata.request_date,
        "vendor_name": quotation.vendor_name,
        "vendor_business_number": quotation.vendor_business_number,
        "vendor_phone_number": quotation.vendor_phone_number,
        "quotation_date": quotation.quotation_date,
        "validity_period": quotation.validity_period,
        "contact_person": quotation.contact_person,
        "supply_amount": quotation.supply_amount,
        "tax_amount": quotation.tax_amount,
        "total_amount": quotation.total_amount,
        "notes": quotation.notes,
    }


def _apply_template_mapping(
    workbook: Workbook,
    quotation: QuotationData,
    approval_metadata: ApprovalMetadata,
    template_mapping: TemplateMapping,
) -> None:
    values = _field_values(quotation, approval_metadata)
    for mapping in template_mapping.scalar_fields:
        if mapping.field_key not in values or mapping.sheet_name not in workbook.sheetnames:
            continue
        workbook[mapping.sheet_name][mapping.cell] = values[mapping.field_key]

    table = template_mapping.item_table
    if not table or not table.sheet_name or table.sheet_name not in workbook.sheetnames:
        return

    ws = workbook[table.sheet_name]
    for row_offset, item in enumerate(quotation.items):
        row = table.start_row + row_offset
        item_values = item.model_dump()
        for field_key, column_letter in table.columns.items():
            if field_key in item_values:
                ws[f"{column_letter}{row}"] = item_values[field_key]


def generate_excel(
    quotation: QuotationData,
    approval_metadata: ApprovalMetadata,
    template_path: str = "",
    output_path: str = "",
    template_mapping: TemplateMapping | None = None,
) -> Path:
    if template_path:
        template = Path(template_path).expanduser()
        if not template.exists():
            raise FileNotFoundError(f"Excel template does not exist: {template}")
        workbook = load_workbook(template)
    else:
        workbook = Workbook()
        _populate_default_approval_sheet(workbook, quotation)

    if template_mapping:
        _apply_template_mapping(workbook, quotation, approval_metadata, template_mapping)

    if template_path:
        _append_summary_sheet(workbook, quotation, approval_metadata)

    destination = build_default_output_path(quotation, output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination
